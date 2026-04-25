"""
parser.py — SOA (Statement of Activities) Excel Parser for St. Andrew Church
Extracts structured financial data from monthly SOA reports.
"""

import re
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime


def parse_soa(file) -> pd.DataFrame:
    """
    Parse a St. Andrew SOA Excel file and return a clean DataFrame.

    Parameters
    ----------
    file : str, Path, or file-like object
        Path to an .xlsx file or an uploaded file object (e.g., from Streamlit).

    Returns
    -------
    pd.DataFrame with columns:
        report_month, church_name, flow_type, category,
        account_code, account_name, current_actual, current_budget,
        ytd_actual, ytd_budget, annual_budget, ytd_last_year
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb["Sheet"]

    # --- 1. Extract metadata from header rows ---
    church_name = _cell(ws, 1, 1) or "St Andrew"
    date_text = _cell(ws, 7, 1) or ""  # e.g., "February 28, 2026"
    report_month = _parse_month(date_text, _cell(ws, 6, 1))

    # --- 2. Walk rows and build records ---
    records = []
    flow_type = None       # "Income" or "Expense"
    category = None        # e.g., "General", "Liturgy", …
    pending_category = None  # for single-line categories like Capital Expense

    max_row = ws.max_row
    for row_idx in range(12, max_row + 1):  # data starts after header row 11
        text = _cell(ws, row_idx, 1)
        if text is None or str(text).strip() == "":
            continue
        text = str(text).strip()

        # --- Section switches ---
        if text == "Income":
            flow_type = "Income"
            continue
        if text == "Expense":
            flow_type = "Expense"
            continue

        # --- Skip totals and summary rows ---
        if "Totals:" in text or text.startswith("Income - Expense"):
            continue
        if text.startswith("Printed:") or text == "Account Shortcut and Description":
            continue
        if text.startswith("Page "):
            continue

        # --- Check if this is a data row (has numbers in the value columns) ---
        vals = _extract_values(ws, row_idx)
        has_values = any(v is not None for v in vals)

        if not has_values:
            # This is a category header (text-only row like "Liturgy", "General", etc.)
            if flow_type:
                category = text
            continue

        # --- Parse account code and name ---
        # Handle single-line categories (Capital Expense, School Subsidy)
        # that have a header row immediately followed by a single data row
        # We already set category when we saw the header.

        account_code, account_name = _parse_account(text)
        if account_code is None:
            # Row has numbers but no parseable account code — skip
            continue

        records.append({
            "report_month": report_month,
            "church_name": church_name.strip(),
            "flow_type": flow_type,
            "category": category,
            "account_code": account_code,
            "account_name": account_name,
            "current_actual": _to_float(vals[0]),
            "current_budget": _to_float(vals[1]),
            "ytd_actual": _to_float(vals[2]),
            "ytd_budget": _to_float(vals[3]),
            "annual_budget": _to_float(vals[4]),
            "ytd_last_year": _to_float(vals[5]),
        })

    wb.close()

    df = pd.DataFrame(records)
    if df.empty:
        return df

    # Ensure numeric columns
    num_cols = [
        "current_actual", "current_budget", "ytd_actual",
        "ytd_budget", "annual_budget", "ytd_last_year",
    ]
    for c in num_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

    return df


# ──────────────────────────────────────────────
#  Helper functions
# ──────────────────────────────────────────────

def _cell(ws, row: int, col: int):
    """Return value of a cell (1-indexed row & col)."""
    return ws.cell(row=row, column=col).value


def _extract_values(ws, row_idx: int) -> tuple:
    """
    Extract the 6 numeric columns from a detail row.
    Returns (current_actual, current_budget, ytd_actual,
             ytd_budget, annual_budget, ytd_last_year).
    Column indices (1-based): 9, 12, 16, 19, 21, 25
    """
    cols = [9, 12, 16, 19, 21, 25]
    return tuple(_cell(ws, row_idx, c) for c in cols)


def _to_float(val) -> float:
    """Convert a cell value to float, handling parenthesized negatives and strings."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("$", "").replace(",", "")
    # Handle accounting-style negatives: (1,234.56)
    m = re.match(r"^\((.+)\)$", s)
    if m:
        try:
            return -float(m.group(1))
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_account(text: str) -> tuple:
    """
    Parse an account line like '4010.000.00.00.00P  Offertory Main'
    into (account_code, account_name).
    Returns (None, None) if it doesn't look like an account line.
    """
    text = text.strip()
    # Standard pattern: code starting with 4 or 5, followed by dots/digits, then name
    # Also handle sub-accounts like "1  Grants/Distributions"
    m = re.match(r"^([45]\d[\d.]+\w*)\s{2,}(.+)$", text)
    if m:
        return m.group(1).strip(), m.group(2).strip()

    # Sub-account pattern: digit(s) followed by double-space and name
    m2 = re.match(r"^(\d+)\s{2,}(.+)$", text)
    if m2:
        return m2.group(1).strip(), m2.group(2).strip()

    return None, None


def _parse_month(date_text: str, title_text: str = None) -> str:
    """
    Parse the report month from date text like 'February 28, 2026'
    or title like 'February 2026 - FINAL'.
    Returns 'YYYY-MM' string.
    """
    # Try full date first
    for fmt in ("%B %d, %Y", "%b %d, %Y"):
        try:
            dt = datetime.strptime(date_text.strip(), fmt)
            return dt.strftime("%Y-%m")
        except (ValueError, AttributeError):
            pass

    # Try title format
    if title_text:
        m = re.match(r"(\w+)\s+(\d{4})", str(title_text))
        if m:
            try:
                dt = datetime.strptime(f"{m.group(1)} 1, {m.group(2)}", "%B %d, %Y")
                return dt.strftime("%Y-%m")
            except ValueError:
                pass

    return "Unknown"


# ──────────────────────────────────────────────
#  CLI test harness
# ──────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    from pathlib import Path

    # Default test files
    test_dir = Path(__file__).parent
    test_files = sorted(test_dir.glob("SOA_*_FINAL.xlsx"))

    if not test_files:
        print("No SOA files found in current directory.")
        sys.exit(1)

    all_dfs = []
    for f in test_files:
        print(f"\n{'='*60}")
        print(f"Parsing: {f.name}")
        print(f"{'='*60}")
        df = parse_soa(f)
        print(f"  Rows extracted: {len(df)}")
        print(f"  Month: {df['report_month'].iloc[0] if len(df) else 'N/A'}")
        print(f"  Flow types: {df['flow_type'].unique().tolist()}")
        print(f"  Categories: {df['category'].unique().tolist()}")

        # Summary by flow_type
        summary = df.groupby("flow_type")["current_actual"].sum()
        for ft, val in summary.items():
            print(f"  {ft} total (current month): ${val:,.2f}")

        all_dfs.append(df)

    combined = pd.concat(all_dfs, ignore_index=True)
    print(f"\n{'='*60}")
    print(f"COMBINED: {len(combined)} rows across {combined['report_month'].nunique()} months")
    print(f"Months: {sorted(combined['report_month'].unique())}")
    print(f"\nSample rows:")
    print(combined[["report_month", "flow_type", "category", "account_code",
                     "account_name", "current_actual", "current_budget"]].head(10).to_string(index=False))
